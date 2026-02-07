"""
Celery tasks for async report generation.
PDF: ReportLab + matplotlib (pure Python, no system dependencies)
Excel: pandas + openpyxl (pure Python)
"""
import os
import sys
import asyncio
import io
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Dict, List, Any, Optional

# Django setup
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent / 'metrq_dj'))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'metriq_site.settings')

import django
django.setup()

from celery import shared_task
from django.utils import timezone
from django.contrib.auth.models import User
from django.db import transaction
from django.core.cache import cache

# ReportLab imports (pure Python)
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image, PageBreak, KeepTogether
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY

# Matplotlib imports (pure Python)
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

# Pandas and openpyxl (pure Python)
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference

from core.models import Report, Article, ProviderLog
from accounts.models import Profile

import structlog

logger = structlog.get_logger(__name__)


# ============================================================================
# DATA SERVICE
# ============================================================================

class ReportDataService:
    """Aggregate data for reports from database."""

    def __init__(self, user_id: str, report_type: str):
        self.user_id = user_id
        self.report_type = report_type
        self.user = User.objects.get(id=user_id)
        self.profile = self.user.profile

        # Calculate date range
        end_date = timezone.now()
        if report_type == 'weekly':
            start_date = end_date - timedelta(days=7)
        else:  # monthly
            start_date = end_date - timedelta(days=30)

        self.start_date = start_date
        self.end_date = end_date

    def get_articles(self) -> List[Article]:
        """Get articles for report period."""
        return Article.objects.filter(
            scraped_at__gte=self.start_date,
            scraped_at__lte=self.end_date
        ).order_by('-scraped_at')

    def get_metrics(self) -> Dict[str, Any]:
        """Calculate key metrics."""
        articles = self.get_articles()

        # Basic counts
        total_articles = len(articles)
        by_language = {}
        for lang in ['zh_cn', 'zh_tw', 'en', 'ru']:
            by_language[lang] = sum(1 for a in articles if a.language == lang)

        # Sentiment analysis
        sentiments = [a.sentiment for a in articles if a.sentiment is not None]
        avg_sentiment = sum(sentiments) / len(sentiments) if sentiments else 0

        # Bias analysis
        biases = [a.bias for a in articles if a.bias is not None]
        avg_bias = sum(biases) / len(biases) if biases else 0.5

        # Entity extraction
        entity_counts = {}
        for article in articles:
            if article.entities:
                import json
                try:
                    entities = json.loads(article.entities) if isinstance(article.entities,
                                                                          str) else article.entities
                    for entity_type, items in entities.items():
                        if isinstance(items, list):
                            for item in items:
                                key = f"{entity_type}:{item}"
                                entity_counts[key] = entity_counts.get(key, 0) + 1
                except:
                    pass

        top_entities = sorted(
            entity_counts.items(),
            key=lambda x: x[1],
            reverse=True
        )[:20]

        # Sentiment timeline (daily averages)
        sentiment_timeline = []
        current_date = self.start_date.date()
        while current_date <= self.end_date.date():
            day_articles = [
                a for a in articles
                if a.scraped_at.date() == current_date and a.sentiment is not None
            ]
            if day_articles:
                avg = sum(a.sentiment for a in day_articles) / len(day_articles)
                sentiment_timeline.append({
                    'date': current_date,
                    'sentiment': avg,
                    'count': len(day_articles)
                })
            current_date += timedelta(days=1)

        return {
            'total_articles': total_articles,
            'by_language': by_language,
            'avg_sentiment': avg_sentiment,
            'avg_bias': avg_bias,
            'top_entities': top_entities,
            'sentiment_timeline': sentiment_timeline,
            'period_start': self.start_date,
            'period_end': self.end_date,
            'user_tier': self.profile.tier
        }


# ============================================================================
# CHART GENERATION (Matplotlib)
# ============================================================================

def create_sentiment_timeline_chart(timeline_data: List[Dict], width=6, height=4, dpi=100):
    """Create sentiment timeline chart using matplotlib."""
    fig, ax = plt.subplots(figsize=(width, height), dpi=dpi)

    if not timeline_data:
        ax.text(0.5, 0.5, 'No data available', ha='center', va='center')
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
    else:
        dates = [item['date'] for item in timeline_data]
        sentiments = [item['sentiment'] for item in timeline_data]
        counts = [item['count'] for item in timeline_data]

        # Plot sentiment line
        ax.plot(dates, sentiments, marker='o', linewidth=2, markersize=4,
                color='steelblue', label='Sentiment')

        # Add zero line
        ax.axhline(y=0, color='gray', linestyle='--', alpha=0.5)

        # Color based on sentiment
        for i, (date, sent) in enumerate(zip(dates, sentiments)):
            color = 'green' if sent > 0.3 else 'red' if sent < -0.3 else 'gray'
            ax.plot(date, sent, 'o', color=color, markersize=4)

        # Formatting
        ax.set_xlabel('Date', fontsize=10)
        ax.set_ylabel('Sentiment Score', fontsize=10)
        ax.set_title('Sentiment Trend Over Time', fontsize=12, fontweight='bold')
        ax.set_ylim(-1, 1)
        ax.grid(True, alpha=0.3)

        # Rotate date labels
        plt.xticks(rotation=45, ha='right')

    plt.tight_layout()

    # Save to buffer
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=dpi, bbox_inches='tight')
    plt.close(fig)
    buffer.seek(0)
    return buffer


def create_language_distribution_chart(by_language: Dict, width=6, height=4, dpi=100):
    """Create language distribution pie chart."""
    fig, ax = plt.subplots(figsize=(width, height), dpi=dpi)

    labels = []
    sizes = []
    colors_map = {'zh_cn': '#ff9999', 'zh_tw': '#66b3ff', 'en': '#99ff99', 'ru': '#ffcc99'}

    for lang, count in by_language.items():
        if count > 0:
            labels.append(lang.replace('_', ' ').upper())
            sizes.append(count)

    if sizes:
        ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90,
               colors=[colors_map.get(l.replace(' ', '_').lower(), 'gray') for l in labels])
        ax.set_title('Articles by Language', fontsize=12, fontweight='bold')
    else:
        ax.text(0.5, 0.5, 'No data available', ha='center', va='center')

    plt.tight_layout()

    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=dpi, bbox_inches='tight')
    plt.close(fig)
    buffer.seek(0)
    return buffer


# ============================================================================
# PDF GENERATION (ReportLab)
# ============================================================================

def generate_pdf_content(metrics: Dict[str, Any]) -> bytes:
    """
    Generate PDF report using ReportLab (pure Python).
    Entirely in memory - no temporary files.
    """
    buffer = io.BytesIO()

    # Create document
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=0.6 * inch,
        leftMargin=0.6 * inch,
        topMargin=0.8 * inch,
        bottomMargin=0.8 * inch
    )

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    heading2_style = ParagraphStyle(
        'CustomHeading2',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#2c5282'),
        spaceAfter=12,
        spaceBefore=12
    )

    normal_style = styles['Normal']
    normal_style.fontSize = 10
    normal_style.leading = 14

    # Build document
    story = []

    # Title page
    story.append(Paragraph("MetrQ Analytics Report", title_style))
    story.append(Spacer(1, 0.3 * inch))

    # Metadata table
    meta_data = [
        ['Report Type:', metrics['report_type'].title()],
        ['Period:', f"{metrics['period_start'].strftime('%Y-%m-%d')} to {metrics['period_end'].strftime('%Y-%m-%d')}"],
        ['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M UTC')],
        ['User Tier:', metrics['user_tier'].title()],
        ['Total Articles:', str(metrics['total_articles'])],
    ]

    meta_table = Table(meta_data, colWidths=[2 * inch, 4 * inch])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.3 * inch))

    # Executive Summary
    story.append(Paragraph("Executive Summary", heading2_style))
    summary_text = f"""
    This report covers <b>{metrics['total_articles']}</b> articles published across 
    <b>{len([v for v in metrics['by_language'].values() if v > 0])}</b> languages. 
    The average sentiment score was <b>{metrics['avg_sentiment']:.2f}</b> 
    (range: -1 negative to +1 positive), with an average bias score of 
    <b>{metrics['avg_bias']:.2f}</b> (range: 0 neutral to 1 highly biased).
    """
    story.append(Paragraph(summary_text, normal_style))
    story.append(Spacer(1, 0.2 * inch))

    # Sentiment Chart
    story.append(Paragraph("Sentiment Analysis", heading2_style))
    story.append(Paragraph(
        "Daily sentiment scores showing the overall tone of coverage over the reporting period.",
        normal_style
    ))
    story.append(Spacer(1, 0.1 * inch))

    # Generate and add chart
    sentiment_chart = create_sentiment_timeline_chart(metrics['sentiment_timeline'])
    story.append(Image(sentiment_chart, width=6 * inch, height=3 * inch))
    story.append(Spacer(1, 0.2 * inch))

    # Language Distribution
    story.append(Paragraph("Language Distribution", heading2_style))
    lang_chart = create_language_distribution_chart(metrics['by_language'])
    story.append(Image(lang_chart, width=4 * inch, height=2.5 * inch))
    story.append(Spacer(1, 0.2 * inch))

    # Language breakdown table
    lang_data = [['Language', 'Article Count', 'Percentage']]
    total = metrics['total_articles'] or 1  # Avoid div by zero
    for lang, count in metrics['by_language'].items():
        pct = (count / total) * 100
        lang_data.append([
            lang.replace('_', ' ').upper(),
            str(count),
            f"{pct:.1f}%"
        ])

    lang_table = Table(lang_data, colWidths=[2 * inch, 2 * inch, 2 * inch])
    lang_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5282')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(lang_table)
    story.append(PageBreak())

    # Top Entities
    story.append(Paragraph("Top Mentioned Entities", heading2_style))
    story.append(Paragraph(
        "Most frequently mentioned persons, organizations, and terms across all articles.",
        normal_style
    ))
    story.append(Spacer(1, 0.1 * inch))

    if metrics['top_entities']:
        entity_data = [['Rank', 'Entity', 'Type', 'Mentions']]
        for i, (entity_key, count) in enumerate(metrics['top_entities'][:15], 1):
            parts = entity_key.split(':', 1)
            entity_type = parts[0].title() if len(parts) > 1 else 'Unknown'
            entity_name = parts[1] if len(parts) > 1 else entity_key
            entity_data.append([str(i), entity_name, entity_type, str(count)])

        entity_table = Table(entity_data, colWidths=[0.5 * inch, 3 * inch, 1.5 * inch, 1 * inch])
        entity_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5282')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(entity_table)
    else:
        story.append(Paragraph("No entity data available for this period.", normal_style))

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================================
# EXCEL GENERATION (Pandas + OpenPyXL)
# ============================================================================

def generate_excel_content(metrics: Dict[str, Any]) -> bytes:
    """
    Generate Excel report using pandas and openpyxl (pure Python).
    Multiple sheets with embedded charts.
    """
    # Create output buffer
    output = io.BytesIO()

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet and create our sheets
    wb.remove(wb.active)

    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "MetrQ Analytics Report"
    ws_summary['A1'].font = Font(size=16, bold=True, color="1F4E79")

    # Metadata
    ws_summary['A3'] = "Report Type:"
    ws_summary['B3'] = metrics['report_type'].title()
    ws_summary['A4'] = "Period Start:"
    ws_summary['B4'] = metrics['period_start'].strftime('%Y-%m-%d')
    ws_summary['A5'] = "Period End:"
    ws_summary['B5'] = metrics['period_end'].strftime('%Y-%m-%d')
    ws_summary['A6'] = "Generated:"
    ws_summary['B6'] = timezone.now().strftime('%Y-%m-%d %H:%M UTC')
    ws_summary['A7'] = "User Tier:"
    ws_summary['B7'] = metrics['user_tier'].title()

    for row in range(3, 8):
        ws_summary[f'A{row}'].font = Font(bold=True)

    # Key Metrics
    ws_summary['A9'] = "Key Metrics"
    ws_summary['A9'].font = Font(size=14, bold=True, color="1F4E79")

    metrics_data = [
        ["Total Articles", metrics['total_articles']],
        ["Average Sentiment", round(metrics['avg_sentiment'], 3)],
        ["Average Bias", round(metrics['avg_bias'], 3)],
        ["Languages Covered", len([v for v in metrics['by_language'].values() if v > 0])],
    ]

    for i, (label, value) in enumerate(metrics_data, start=10):
        ws_summary[f'A{i}'] = label
        ws_summary[f'B{i}'] = value
        ws_summary[f'A{i}'].font = Font(bold=True)

    # Language Distribution
    ws_summary['A15'] = "Articles by Language"
    ws_summary['A15'].font = Font(size=14, bold=True, color="1F4E79")

    lang_start = 16
    ws_summary[f'A{lang_start}'] = "Language"
    ws_summary[f'B{lang_start}'] = "Count"
    ws_summary[f'C{lang_start}'] = "Percentage"

    for cell in [ws_summary[f'A{lang_start}'], ws_summary[f'B{lang_start}'],
                 ws_summary[f'C{lang_start}']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    total = metrics['total_articles'] or 1
    row = lang_start + 1
    for lang, count in metrics['by_language'].items():
        ws_summary[f'A{row}'] = lang.replace('_', ' ').upper()
        ws_summary[f'B{row}'] = count
        ws_summary[f'C{row}'] = f"{(count / total) * 100:.1f}%"
        row += 1

    # Add bar chart for languages
    chart = BarChart()
    chart.type = "col"
    chart.title = "Articles by Language"
    chart.x_axis.title = "Language"
    chart.y_axis.title = "Count"

    data = Reference(ws_summary, min_col=2, min_row=lang_start, max_row=row - 1)
    cats = Reference(ws_summary, min_col=1, min_row=lang_start + 1, max_row=row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws_summary.add_chart(chart, "E9")

    # Sheet 2: Sentiment Timeline
    ws_sentiment = wb.create_sheet("Sentiment Trend")

    ws_sentiment['A1'] = "Date"
    ws_sentiment['B1'] = "Articles"
    ws_sentiment['C1'] = "Avg Sentiment"

    for cell in [ws_sentiment['A1'], ws_sentiment['B1'], ws_sentiment['C1']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for i, item in enumerate(metrics['sentiment_timeline'], start=2):
        ws_sentiment[f'A{i}'] = item['date'].strftime('%Y-%m-%d')
        ws_sentiment[f'B{i}'] = item['count']
        ws_sentiment[f'C{i}'] = round(item['sentiment'], 3)

    # Add line chart for sentiment
    line_chart = LineChart()
    line_chart.title = "Sentiment Trend Over Time"
    line_chart.x_axis.title = "Date"
    line_chart.y_axis.title = "Sentiment Score"
    line_chart.y_axis.scaling.min = -1
    line_chart.y_axis.scaling.max = 1

    dates = Reference(ws_sentiment, min_col=1, min_row=2, max_row=len(metrics['sentiment_timeline']) + 1)
    sentiments = Reference(ws_sentiment, min_col=3, min_row=1,
                           max_row=len(metrics['sentiment_timeline']) + 1)
    line_chart.add_data(sentiments, titles_from_data=True)
    line_chart.set_categories(dates)
    ws_sentiment.add_chart(line_chart, "E2")

    # Sheet 3: Top Entities
    ws_entities = wb.create_sheet("Top Entities")

    ws_entities['A1'] = "Rank"
    ws_entities['B1'] = "Entity"
    ws_entities['C1'] = "Type"
    ws_entities['D1'] = "Mentions"

    for cell in [ws_entities['A1'], ws_entities['B1'], ws_entities['C1'], ws_entities['D1']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for i, (entity_key, count) in enumerate(metrics['top_entities'][:50], start=2):
        parts = entity_key.split(':', 1)
        entity_type = parts[0].title() if len(parts) > 1 else 'Unknown'
        entity_name = parts[1] if len(parts) > 1 else entity_key

        ws_entities[f'A{i}'] = i - 1
        ws_entities[f'B{i}'] = entity_name
        ws_entities[f'C{i}'] = entity_type
        ws_entities[f'D{i}'] = count

    # Auto-adjust column widths
    for ws in [ws_summary, ws_sentiment, ws_entities]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to buffer
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================================
# CELERY TASKS
# ============================================================================

@shared_task(bind=True, max_retries=3)
def generate_pdf_report(self, job_id: str, user_id: str, report_type: str):
    """
    Async task to generate PDF report using ReportLab + matplotlib.
    Wrapped in asyncio.to_thread() for async compatibility.
    """
    try:
        # Update status
        _update_report_progress(job_id, 10, 'processing')

        # Get data
        data_service = ReportDataService(user_id, report_type)
        metrics = data_service.get_metrics()
        metrics['report_type'] = report_type

        _update_report_progress(job_id, 40, 'processing')

        # Generate PDF in thread (CPU-bound)
        loop = asyncio.get_event_loop()
        pdf_bytes = loop.run_in_executor(
            None,
            generate_pdf_content,
            metrics
        )

        _update_report_progress(job_id, 80, 'processing')

        # Save to database
        pdf_data = pdf_bytes.result() if asyncio.isfuture(pdf_bytes) else pdf_bytes

        with transaction.atomic():
            report = Report.objects.get(id=job_id)
            report.file_blob = pdf_data
            report.status = 'done'
            report.completed_at = timezone.now()
            report.save()

        _update_report_progress(job_id, 100, 'done')

        logger.info("pdf_report_generated",
                    job_id=job_id,
                    user_id=user_id,
                    size_bytes=len(pdf_data))

        return {"job_id": job_id, "status": "done", "size": len(pdf_data)}

    except Exception as e:
        logger.error("pdf_generation_failed", job_id=job_id, error=str(e))

        # Update status to failed
        try:
            report = Report.objects.get(id=job_id)
            report.status = 'failed'
            report.error_message = str(e)[:500]
            report.save()
        except:
            pass

        # Retry
        raise self.retry(exc=e, countdown=60)


@shared_task(bind=True, max_retries=3)
def generate_excel_report(self, job_id: str, user_id: str, report_type: str):
    """
    Async task to generate Excel report using pandas + openpyxl.
    Faster than PDF generation.
    """
    try:
        _update_report_progress(job_id, 10, 'processing')

        data_service = ReportDataService(user_id, report_type)
        metrics = data_service.get_metrics()
        metrics['report_type'] = report_type

        _update_report_progress(job_id, 50, 'processing')

        # Generate Excel in thread
        loop = asyncio.get_event_loop()
        excel_bytes = loop.run_in_executor(
            None,
            generate_excel_content,
            metrics
        )

        _update_report_progress(job_id, 90, 'processing')

        excel_data = excel_bytes.result() if asyncio.isfuture(excel_bytes) else excel_bytes

        with transaction.atomic():
            report = Report.objects.get(id=job_id)
            report.file_blob = excel_data
            report.status = 'done'
            report.completed_at = timezone.now()
            report.save()

        _update_report_progress(job_id, 100, 'done')

        logger.info("excel_report_generated",
                    job_id=job_id,
                    user_id=user_id,
                    size_bytes=len(excel_data))

        return {"job_id": job_id, "status": "done", "size": len(excel_data)}

    except Exception as e:
        logger.error("excel_generation_failed", job_id=job_id, error=str(e))

        try:
            report = Report.objects.get(id=job_id)
            report.status = 'failed'
            report.error_message = str(e)[:500]
            report.save()
        except:
            pass

        raise self.retry(exc=e, countdown=60)


def _update_report_progress(job_id: str, progress: int, status: str):
    """Update report progress in Redis cache."""
    cache.set(f"metrq:report:{job_id}:progress", progress, timeout=3600)
    if status:
        cache.set(f"metrq:report:{job_id}:status", status, timeout=3600)


@shared_task
def cleanup_old_reports():
    """
    Daily cleanup of old report files to save disk space.
    Keep reports for 30 days.
    """
    cutoff = timezone.now() - timedelta(days=30)
    old_reports = Report.objects.filter(
        created_at__lt=cutoff,
        status='done'
    )

    count = 0
    for report in old_reports:
        report.file_blob = None  # Clear blob but keep metadata
        report.save()
        count += 1

    logger.info("old_reports_cleaned_up", count=count)
    return f"Cleared {count} old reports"
